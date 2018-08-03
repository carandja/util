import sys
from openpyxl import load_workbook

def run():
    print('<html><table>')
    wb = load_workbook(filename = sys.argv[1])
    ws = wb.active
    rr = cc = 1
    formatrow = head
    while ws.cell(row = rr, column = cc).value != None:
        items = []
        while ws.cell(row = rr, column = cc).value != None:
            items.append(ws.cell(row = rr, column = cc).value)
            cc += 1
        print(formatrow(items))
        formatrow = row
        rr += 1
        cc = 1
    print('</table></html>')

def head(items):
    return makerow('h', items)

def row(items):
    return makerow('d', items)

def makerow(rowtype, items):
    tagstart = '<t{}>'.format(rowtype)
    tagend = '</t{}>'.format(rowtype)
    return '<tr>{}</tr>'.format(''.join([tagstart + str(i) + tagend for i in items]))
run()
